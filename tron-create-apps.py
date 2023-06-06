# tron-create apps is a python script that reads a spreadsheet, creates a json object/file and then uses the Private
# Apps API to post this json data and create a new private app.
# 
# It currently uses a Netskope APIv2 endpoint and needs Read/Write access:
# 
# '/api/v2/steering/apps/private'
#
# It requires 3 arguments in order:  tenant name, private app API key, excel filename
# excel file must end in .xlsx
#
# The column names in the excel file must be app_id, app_name, host, use_publisher_dns, port/protocol, publisher_name/publisher_id, \ 
# clientless_access, trust_self_signed_certs
#
#
# This was written for Netskope release 101 originally, python 3.10 
# 
#
#
####################################################################################



import sys
import logging
import requests
import json
from openpyxl import load_workbook

# set logging level
logging.basicConfig(level=logging.WARNING)


####### Check for correct arguments and deliver help message

if len(sys.argv) > 1 and sys.argv[1] == '--help':
        print('Full tenant hostname, API token, destination excel filename required.  In order \n')
        print('Example:    python3 tron-create-apps hostname.goskope.com a458ef832xs389 myexcel.xlsx')
        exit()


if len(sys.argv) < 3:
    print("Error: This program requires 3 arguments: Full tenant hostname, API token, excel filename")
    print('Example:    python3 tron-create-apps hostname.goskope.com a458ef832xs389 myexcel.xlsx')
    sys.exit(1)


arg1=sys.argv[1]
arg2=sys.argv[2]
arg3=sys.argv[3]



if len([arg1, arg2, arg3]) < 3:
        raise Exception("This function requires 3 arguments")

# setup input arguments
tenant = str(arg1)
API_token = str(arg2)
excelfile = str(arg3)

# setup URL and HTTP headers
API_URL = '/api/v2/steering/apps/private'
appurlrequest = 'http://' + tenant + API_URL
httpheaders = {'Content-type': 'application/json', 'Netskope-Api-Token': API_token }

# Load the workbook
wb = load_workbook(filename=excelfile)

# Select the worksheet
ws = wb['Sheet1']

# Get the column names from the first row
header = [cell.value for cell in ws[1]]

# Rename a couple of header column names to match json required.
header[4] = 'protocols'
header[5] = 'publishers'

# Get the column indices for the "protocols" and "publishers" columns, they will need to be modified
port_protocol_index = header.index('protocols') + 1
publisher_index = header.index('publishers') + 1

# Create a list to store the data
data = []

# Iterate over the rows (skipping the first row)
for row in ws.iter_rows(min_row=2):

    # Create a dictionary to store the data for this row
    row_data = {}

    # Iterate over the cells in the row
    for cell in row:

        # Get the column name for this cell
        column_name = header[cell.column - 1]

        # skip the app_id column since its not needed
        if cell.column == 1:
            continue

        # Get the value of the cell
        value = cell.value

        # If this is the "protocols" split the value into an array
        if cell.column == port_protocol_index:
            # create an array
            result = []
            
            # split the values at the commas
            protocol_value = value.split(',')

            # loop through the split values
            for item in protocol_value:
                #split them into port/protocol
                port, protocol_type = item.split('/')

                # format them for the API
                tempvalue = {'type': protocol_type, 'port': port }
                
                # append them to the array for json
                result.append(tempvalue)
            # Asign then entire array to value    
            value = result
                
            

        # if this is the "publishers" split the value into an array
        if cell.column == publisher_index:
            
            # create an array
            result = []

            # split the values at the commas
            publisher_value = value.split(',')

            # loop through the split values
            for item in publisher_value:

                # split them into publisher name and ID
                publisher_name, publisher_id = item.split('/')

                # format for the API
                tempvalue = {'publisher_id': publisher_id, 'publisher_name': publisher_name, }

                # Append them to the array for json
                result.append(tempvalue)
            # Assing the entire array to value    
            value = result

        # Add the data from value to the dictionary (of json)
        row_data[column_name] = value
        
    # Add the dictionary to the list
    data.append(row_data)

    # create a json document of just the row for the API call
    jsonrow_data = json.dumps(row_data, indent=4)
    print(jsonrow_data)

    # POST the data and print the server response
    response = requests.post(appurlrequest, jsonrow_data, headers=httpheaders)
    print(response)

# Convert the data to JSON
json_data = json.dumps(data, indent=4)

# Print the JSON data
print(json_data)

with open( 'temp-tron-create-apps-json.json', 'w') as f:
        json.dump(data, f, indent=4,)




